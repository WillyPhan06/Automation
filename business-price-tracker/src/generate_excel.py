from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import pandas as pd
from pathlib import Path

def format_excel(input_csv, output_excel):
    df = pd.read_csv(input_csv)

    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    # Write headers
    for col_num, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_num, value=col_name)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4F81BD", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # Write data
    for row_num, row in enumerate(df.values, start=2):
        for col_num, value in enumerate(row, start=1):
            ws.cell(row=row_num, column=col_num, value=value)

    # Auto-fit columns
    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 2

    wb.save(output_excel)
    print(f"✅ Excel report saved: {output_excel}")

if __name__ == "__main__":
    format_excel("data/cleaned/cleaned_data.csv", "reports/prices_report.xlsx")
